import csv
import io
import unittest
import zipfile
from unittest.mock import patch

import openpyxl

from purchase_service import (
    PurchaseRow,
    build_output_excel,
    build_template_csv,
    build_zip,
    find_new_barcodes,
    import_new_barcodes,
    parse_purchase_excel,
)


def _make_excel(data_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["条码", "col2", "价格", "col4", "col5", "数量", "col7"])
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestPurchaseRow(unittest.TestCase):
    def test_formatted_four_decimal(self):
        row = PurchaseRow(
            barcode="1234567890123", price_raw="9.48", price=9.48, quantity=144, price_flagged=False
        )
        self.assertEqual(row.formatted(), "1234567890123,9.4800,,144")

    def test_formatted_pads_whole_number_to_four_decimal(self):
        row = PurchaseRow(
            barcode="1234567890123", price_raw="12", price=12.0, quantity=36, price_flagged=False
        )
        self.assertEqual(row.formatted(), "1234567890123,12.0000,,36")

    def test_to_dict_has_expected_keys(self):
        row = PurchaseRow(
            barcode="1111", price_raw="5.0", price=5.0, quantity=10, price_flagged=False
        )
        d = row.to_dict()
        self.assertEqual(d["barcode"], "1111")
        self.assertAlmostEqual(d["price"], 5.0)
        self.assertEqual(d["quantity"], 10)
        self.assertFalse(d["price_flagged"])
        self.assertEqual(d["formatted"], "1111,5.0000,,10")


class TestParsePurchaseExcel(unittest.TestCase):
    def test_parses_basic_row(self):
        data = _make_excel([["1234567890123", "x", 9.48, "x", "x", 144, "x"]])
        rows = parse_purchase_excel(data)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].barcode, "1234567890123")
        self.assertAlmostEqual(rows[0].price, 9.48)
        self.assertEqual(rows[0].quantity, 144)
        self.assertFalse(rows[0].price_flagged)

    def test_rounds_price_beyond_four_decimals(self):
        data = _make_excel([["1234567890123", "x", 9.48125, "x", "x", 10, "x"]])
        rows = parse_purchase_excel(data)
        self.assertFalse(rows[0].price_flagged)
        self.assertAlmostEqual(rows[0].price, 9.4813)

    def test_keeps_four_decimals_as_is(self):
        data = _make_excel([["1234567890123", "x", 9.4812, "x", "x", 10, "x"]])
        rows = parse_purchase_excel(data)
        self.assertFalse(rows[0].price_flagged)
        self.assertAlmostEqual(rows[0].price, 9.4812)

    def test_numeric_barcode_with_empty_row_has_no_decimal(self):
        data = _make_excel(
            [
                [6901234567890, "x", 9.48, "x", "x", 10, "x"],
                [None, "x", 9.48, "x", "x", 5, "x"],
            ]
        )
        rows = parse_purchase_excel(data)
        self.assertEqual(rows[0].barcode, "6901234567890")

    def test_parses_multiple_data_rows_skips_header(self):
        data = _make_excel(
            [
                ["BC1", "x", 1.0, "x", "x", 5, "x"],
                ["BC2", "x", 2.0, "x", "x", 3, "x"],
            ]
        )
        rows = parse_purchase_excel(data)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].barcode, "BC1")
        self.assertEqual(rows[1].barcode, "BC2")


class TestBuildOutputExcel(unittest.TestCase):
    def test_appends_header_and_data(self):
        file_bytes = _make_excel([["BC1", "x", 9.48, "x", "x", 10, "x"]])
        rows_data = [{"formatted": "BC1,9.48,,10"}]
        result = build_output_excel(file_bytes, rows_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=8).value, "导入信息")
        self.assertEqual(ws.cell(row=2, column=8).value, "BC1,9.48,,10")

    def test_original_data_preserved(self):
        file_bytes = _make_excel([["BC1", "x", 9.48, "x", "x", 10, "x"]])
        result = build_output_excel(file_bytes, [{"formatted": "BC1,9.48,,10"}])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "BC1")
        self.assertAlmostEqual(ws.cell(row=2, column=3).value, 9.48)


class TestFindNewBarcodes(unittest.TestCase):
    def _row(self, barcode):
        return PurchaseRow(
            barcode=barcode, price_raw="1", price=1.0, quantity=1, price_flagged=False
        )

    def test_all_new_when_system_empty(self):
        rows = [self._row("A"), self._row("B")]
        self.assertEqual(find_new_barcodes(rows, set()), ["A", "B"])

    def test_returns_only_rows_not_in_system(self):
        rows = [self._row("A"), self._row("B"), self._row("C")]
        self.assertEqual(find_new_barcodes(rows, {"B"}), ["A", "C"])

    def test_dedupes_preserving_order(self):
        rows = [self._row("A"), self._row("A"), self._row("B"), self._row("A")]
        self.assertEqual(find_new_barcodes(rows, set()), ["A", "B"])

    def test_empty_rows_returns_empty(self):
        self.assertEqual(find_new_barcodes([], {"X"}), [])


class TestImportNewBarcodes(unittest.TestCase):
    def _entry(self, barcode):
        return {"barcode": barcode}

    @patch("purchase_service.stockpile_db.insert_or_update")
    def test_inserts_each_entry(self, mock_insert):
        entries = [self._entry("BC1"), self._entry("BC2")]
        count = import_new_barcodes(entries)
        self.assertEqual(count, 2)
        self.assertEqual(mock_insert.call_count, 2)
        mock_insert.assert_any_call(
            barcode="BC1", model="BC1", location="", source="purchase_import"
        )
        mock_insert.assert_any_call(
            barcode="BC2", model="BC2", location="", source="purchase_import"
        )

    @patch("purchase_service.stockpile_db.insert_or_update")
    def test_skips_empty_barcode(self, mock_insert):
        entries = [self._entry("BC1"), self._entry(""), self._entry("BC2")]
        count = import_new_barcodes(entries)
        self.assertEqual(count, 2)

    @patch("purchase_service.stockpile_db.insert_or_update")
    def test_skips_whitespace_only(self, mock_insert):
        entries = [self._entry("  ")]
        count = import_new_barcodes(entries)
        self.assertEqual(count, 0)
        mock_insert.assert_not_called()


class TestBuildTemplateCsv(unittest.TestCase):
    def test_single_entry_has_correct_indices(self):
        entries = [
            {
                "barcode": "1234567890123",
                "name": "测试品",
                "invoice_name": "发票名",
                "supplier_id": "S01",
                "supplier_name": "某供应商",
            }
        ]
        out = build_template_csv(entries)
        text = out.decode("gbk")
        lines = text.splitlines()
        self.assertEqual(len(lines), 2)
        fields = next(csv.reader([lines[1]]))
        self.assertEqual(fields[0], "1234567890123")
        self.assertEqual(fields[1], "1234567890123")
        self.assertEqual(fields[3], "测试品")
        self.assertEqual(fields[4], "发票名")
        self.assertEqual(fields[10], "1234567890123")
        self.assertEqual(fields[38], "S01")
        self.assertEqual(fields[39], "某供应商")
        for i in [2, 5, 6, 7, 8, 9, 11, 12, 37]:
            self.assertEqual(fields[i], "")

    def test_column_count_matches_header(self):
        out = build_template_csv(
            [
                {
                    "barcode": "X",
                    "name": "Y",
                    "invoice_name": "Z",
                    "supplier_id": "S",
                    "supplier_name": "N",
                }
            ]
        )
        lines = out.decode("gbk").splitlines()
        header_fields = next(csv.reader([lines[0]]))
        data_fields = next(csv.reader([lines[1]]))
        self.assertEqual(len(data_fields), len(header_fields))

    def test_multiple_entries(self):
        entries = [
            {
                "barcode": "A",
                "name": "品A",
                "invoice_name": "发A",
                "supplier_id": "S1",
                "supplier_name": "N1",
            },
            {
                "barcode": "B",
                "name": "品B",
                "invoice_name": "发B",
                "supplier_id": "S1",
                "supplier_name": "N1",
            },
        ]
        out = build_template_csv(entries)
        lines = out.decode("gbk").splitlines()
        self.assertEqual(len(lines), 3)


class TestBuildZip(unittest.TestCase):
    def test_xlsx_only_when_template_none(self):
        out = build_zip(b"fake-xlsx", None, "20260418")
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            names = zf.namelist()
        self.assertEqual(names, ["采购订单20260418.xlsx"])

    def test_both_files_when_template_provided(self):
        out = build_zip(b"fake-xlsx", b"fake-csv", "20260418")
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            names = sorted(zf.namelist())
        self.assertEqual(names, sorted(["采购订单20260418.xlsx", "产品信息导入模板.csv"]))

    def test_xlsx_content_preserved(self):
        out = build_zip(b"HELLO", None, "20260101")
        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            self.assertEqual(zf.read("采购订单20260101.xlsx"), b"HELLO")
