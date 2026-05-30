import io
import unittest
from unittest import mock

import openpyxl

from app.services import pricing_sheet


def _eur() -> str:
    return '#,##0.00 "€"'


class TestMedianAndTarget(unittest.TestCase):
    def test_median_odd_even_empty(self):
        self.assertAlmostEqual(pricing_sheet._median([27.0, 35.0, 30.0]), 30.0)
        self.assertAlmostEqual(pricing_sheet._median([27.0, 35.0]), 31.0)
        self.assertIsNone(pricing_sheet._median([]))
        self.assertIsNone(pricing_sheet._median([None, None]))
        self.assertAlmostEqual(pricing_sheet._median([10.0, None, 20.0]), 15.0)

    def test_target_margin_filters_supplier_and_nones(self):
        summary = [
            {"barcode": "A", "supplier_id": "S1", "margin_pct": 35.0},
            {"barcode": "B", "supplier_id": "S1", "margin_pct": 27.0},
            {"barcode": "C", "supplier_id": "S1", "margin_pct": None},
            {"barcode": "D", "supplier_id": "S2", "margin_pct": 99.0},
        ]
        res = pricing_sheet.compute_target_margin_pct(summary, "S1")
        self.assertEqual(res["n_samples"], 2)
        self.assertAlmostEqual(res["median"], 31.0)

    def test_target_margin_no_supplier_or_no_data(self):
        self.assertEqual(pricing_sheet.compute_target_margin_pct([], ""),
                         {"median": None, "n_samples": 0})
        self.assertEqual(
            pricing_sheet.compute_target_margin_pct(
                [{"barcode": "A", "supplier_id": "S9", "margin_pct": None}], "S9"),
            {"median": None, "n_samples": 0})


class TestDetectChanged(unittest.TestCase):
    def test_changed_skipped_and_unknown(self):
        rows = [
            {"barcode": "OLD1", "price": 1.05, "quantity": 50},
            {"barcode": "OLD2", "price": 2.00, "quantity": 10},
            {"barcode": "OLD3", "price": 3.00, "quantity": 5},
            {"barcode": "NEW1", "price": 9.00, "quantity": 1},
        ]
        baselines = {"OLD1": 1.10, "OLD2": 2.00, "OLD3": None}
        changed, skipped = pricing_sheet.detect_changed_old(rows, baselines)
        self.assertEqual(skipped, 1)
        self.assertEqual([c["barcode"] for c in changed], ["OLD1"])
        self.assertAlmostEqual(changed[0]["old_price"], 1.10)
        self.assertAlmostEqual(changed[0]["new_price"], 1.05)
        self.assertEqual(changed[0]["quantity"], 50)

    def test_duplicate_barcode_deduped(self):
        rows = [
            {"barcode": "OLD1", "price": 1.05, "quantity": 50},
            {"barcode": "OLD1", "price": 1.06, "quantity": 10},  # 第二次出现，应被去重
        ]
        changed, skipped = pricing_sheet.detect_changed_old(rows, {"OLD1": 1.10})
        self.assertEqual(len(changed), 1)
        self.assertAlmostEqual(changed[0]["new_price"], 1.05)  # 第一条胜出


class TestBuildItems(unittest.TestCase):
    def test_new_and_changed_assembled(self):
        rows = [
            {"barcode": "NEW1", "price": 5.15, "quantity": 100},
            {"barcode": "OLD1", "price": 1.05, "quantity": 50},
        ]
        new_entries = [{"barcode": "NEW1", "name": "不锈钢勺", "invoice_name": "spoon"}]
        products = {  # 主档已知（=老品），含基准/现售价/品名
            "OLD1": {"name_zh": "陶瓷碗", "sale_price": 1.98, "last_purchase_unit_price": 1.10},
        }
        summary_by_bc = {"OLD1": {"urgency_score": 78.0}}
        out = pricing_sheet.build_pricing_items(new_entries, rows, products, summary_by_bc)
        self.assertEqual(out["skipped_no_baseline"], 0)
        self.assertEqual(len(out["new"]), 1)
        self.assertEqual(len(out["changed"]), 1)

        n = out["new"][0]
        self.assertEqual((n["section"], n["barcode"], n["name_zh"]), ("new", "NEW1", "不锈钢勺"))
        self.assertEqual(n["quantity"], 100)
        self.assertAlmostEqual(n["new_price"], 5.15)
        self.assertIsNone(n["old_price"])
        self.assertIsNone(n["sale_price"])
        self.assertIsNone(n["urgency"])

        c = out["changed"][0]
        self.assertEqual((c["section"], c["barcode"], c["name_zh"]), ("changed", "OLD1", "陶瓷碗"))
        self.assertAlmostEqual(c["old_price"], 1.10)
        self.assertAlmostEqual(c["new_price"], 1.05)
        self.assertAlmostEqual(c["sale_price"], 1.98)
        self.assertAlmostEqual(c["urgency"], 78.0)

    def test_phantom_no_history_surfaces_as_new(self):
        """在主档但完全无历史（无历史进价 且 无销售价格）的'幽灵'产品 → 当新品进表；
        有 sale_price 的真老品(缺进价)不被误判，仍跳过(计入 skipped_no_baseline)。"""
        rows = [
            {"barcode": "GHOST", "price": 3.20, "quantity": 8},     # 幽灵：无进价 无售价
            {"barcode": "REALOLD", "price": 2.00, "quantity": 4},   # 真老品：无进价 但有售价
        ]
        products = {
            "GHOST": {"name_zh": "测试误入品", "sale_price": None, "last_purchase_unit_price": None},
            "REALOLD": {"name_zh": "老货", "sale_price": 5.0, "last_purchase_unit_price": None},
        }
        out = pricing_sheet.build_pricing_items([], rows, products, {})
        new_bcs = {it["barcode"] for it in out["new"]}
        self.assertIn("GHOST", new_bcs)        # 幽灵进新品段
        self.assertNotIn("REALOLD", new_bcs)   # 真老品(有售价)不进新品段
        ghost = next(it for it in out["new"] if it["barcode"] == "GHOST")
        self.assertEqual(ghost["section"], "new")
        self.assertEqual(ghost["name_zh"], "测试误入品")    # 用主档品名
        self.assertAlmostEqual(ghost["new_price"], 3.20)    # 用上传新价
        self.assertEqual(ghost["quantity"], 8)
        self.assertIsNone(ghost["old_price"])
        self.assertIsNone(ghost["sale_price"])
        # REALOLD 仍未列入：计入 skipped，且不在 changed 段；GHOST 不计入 skipped
        self.assertEqual(out["skipped_no_baseline"], 1)
        self.assertNotIn("REALOLD", {it["barcode"] for it in out["changed"]})


class TestBuildXlsx(unittest.TestCase):
    def _items(self):
        return {
            "new": [{"section": "new", "barcode": "NEW1", "name_zh": "勺",
                     "quantity": 100, "old_price": None, "new_price": 5.15,
                     "sale_price": None, "urgency": None}],
            "changed": [{"section": "changed", "barcode": "OLD1", "name_zh": "碗",
                         "quantity": 50, "old_price": 1.10, "new_price": 1.05,
                         "sale_price": 1.98, "urgency": 78.0}],
        }

    def test_layout_formulas_and_format(self):
        data = pricing_sheet.build_pricing_xlsx(self._items(), 0.30, "雅典XX贸易", "20260529")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        # B2 = 目标利润率小数 + 百分比格式
        self.assertAlmostEqual(ws["B2"].value, 0.30)
        self.assertEqual(ws["B2"].number_format, "0.00%")
        # 表头行（第 4 行）新列序
        expected_headers = [
            "图片", "条码", "中文品名", "数量", "旧进价", "新进价", "现售价",
            "老利润率", "推荐新售价", "修改", "改后利润率", "热度",
            "建议批发价", "建议利润率",
        ]
        for idx, name in enumerate(expected_headers, start=1):
            self.assertEqual(ws.cell(row=4, column=idx).value, name)
        # 收集所有单元格文本，确认两个分段标题存在
        texts = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
        self.assertIn("◆ 新产品", texts)
        self.assertIn("◆ 调价老产品", texts)

        # ---- 新产品行：四列同数据（老利润率镜像建议利润率、推荐新售价镜像建议批发价）----
        new_r = next(r for r in range(1, ws.max_row + 1)
                     if ws.cell(row=r, column=2).value == "NEW1")
        self.assertEqual(ws.cell(row=new_r, column=6).value, 5.15)          # 新进价
        self.assertIsNone(ws.cell(row=new_r, column=5).value)              # 旧进价空
        self.assertIsNone(ws.cell(row=new_r, column=7).value)             # 现售价空
        self.assertEqual(ws.cell(row=new_r, column=12).value, "新品")      # 热度（移到第 12 列）
        # 老利润率(8) 镜像建议利润率(N=14)；推荐新售价(9) 镜像建议批发价(M=13)
        self.assertEqual(str(ws.cell(row=new_r, column=8).value), f"=N{new_r}")
        self.assertEqual(str(ws.cell(row=new_r, column=9).value), f"=M{new_r}")
        # 建议批发价(13) 引用 $B$2；建议利润率(14) 引用 M
        self.assertIn("$B$2", str(ws.cell(row=new_r, column=13).value))
        self.assertIn(f"M{new_r}", str(ws.cell(row=new_r, column=14).value))
        self.assertEqual(ws.cell(row=new_r, column=6).number_format, _eur())

        # ---- 调价老品行：真实老利润率 + 推荐新售价 ----
        old_r = next(r for r in range(1, ws.max_row + 1)
                     if ws.cell(row=r, column=2).value == "OLD1")
        self.assertEqual(ws.cell(row=old_r, column=5).value, 1.10)        # 旧进价
        self.assertEqual(ws.cell(row=old_r, column=7).value, 1.98)        # 现售价
        self.assertEqual(ws.cell(row=old_r, column=12).value, 78.0)       # 热度数值
        # 老利润率(8) = (现售价G - 旧进价E)/现售价G —— 关键：用旧进价 E，不是新进价 F
        om = str(ws.cell(row=old_r, column=8).value)
        self.assertIn(f"E{old_r}", om)
        self.assertIn(f"G{old_r}", om)
        self.assertNotIn(f"F{old_r}", om)
        # 推荐新售价(9) = 新进价F / (1 - 老利润率H)
        rs = str(ws.cell(row=old_r, column=9).value)
        self.assertIn(f"F{old_r}", rs)
        self.assertIn(f"1-H{old_r}", rs)
        # 建议批发价(13)/建议利润率(14) 老品也填（参考列）
        self.assertIn("$B$2", str(ws.cell(row=old_r, column=13).value))
        self.assertIn(f"M{old_r}", str(ws.cell(row=old_r, column=14).value))
        # 改后利润率(11) = (修改J - 新进价F)/修改J，两区块都填
        k_old = str(ws.cell(row=old_r, column=11).value)
        self.assertIn(f"J{old_r}", k_old)
        self.assertIn(f"F{old_r}", k_old)
        # 数字格式锁定：价格列 EUR、利润率列 百分比
        for col in (5, 7, 9, 13):   # 旧进价/现售价/推荐新售价/建议批发价
            self.assertEqual(ws.cell(row=old_r, column=col).number_format, _eur())
        for col in (8, 11, 14):     # 老利润率/改后利润率/建议利润率
            self.assertEqual(ws.cell(row=old_r, column=col).number_format, "0.00%")
        # 冻结表头（A5）
        self.assertEqual(ws.freeze_panes, "A5")
        # 图片列空 + 行高加高
        self.assertIsNone(ws.cell(row=new_r, column=1).value)
        self.assertIsNotNone(ws.row_dimensions[new_r].height)
        self.assertGreaterEqual(ws.row_dimensions[new_r].height, 60)

    def test_changed_without_saleprice_falls_back_to_mirror(self):
        """调价老品缺现售价 → 老利润率算不出 → 回退镜像（=N / =M），避免表里空洞。"""
        items = {
            "new": [],
            "changed": [{"section": "changed", "barcode": "OLD9", "name_zh": "无售价",
                         "quantity": 5, "old_price": 2.0, "new_price": 2.5,
                         "sale_price": None, "urgency": 10.0}],
        }
        data = pricing_sheet.build_pricing_xlsx(items, 0.30, "S", "20260529")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        r = next(rr for rr in range(1, ws.max_row + 1)
                 if ws.cell(row=rr, column=2).value == "OLD9")
        self.assertEqual(str(ws.cell(row=r, column=8).value), f"=N{r}")   # 老利润率←建议利润率
        self.assertEqual(str(ws.cell(row=r, column=9).value), f"=M{r}")   # 推荐新售价←建议批发价
        self.assertIsNone(ws.cell(row=r, column=7).value)                 # 现售价列空
        self.assertEqual(ws.cell(row=r, column=5).value, 2.0)             # 旧进价仍在
        self.assertEqual(ws.cell(row=r, column=12).value, 10.0)           # 热度数值

    def test_free_goods_zero_oldprice_falls_back_to_mirror(self):
        """调价老品 old_price=0（free goods/脏数据）→ 算不出老利润率 → 回退镜像，不留空格。"""
        items = {
            "new": [],
            "changed": [{"section": "changed", "barcode": "FREE1", "name_zh": "赠品",
                         "quantity": 3, "old_price": 0.0, "new_price": 1.2,
                         "sale_price": 2.0, "urgency": 5.0}],
        }
        data = pricing_sheet.build_pricing_xlsx(items, 0.30, "S", "20260530")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        r = next(rr for rr in range(1, ws.max_row + 1)
                 if ws.cell(row=rr, column=2).value == "FREE1")
        self.assertEqual(str(ws.cell(row=r, column=8).value), f"=N{r}")   # 老利润率←镜像
        self.assertEqual(str(ws.cell(row=r, column=9).value), f"=M{r}")   # 推荐新售价←镜像


    def test_name_formula_injection_neutralized(self):
        items = {
            "new": [{"section": "new", "barcode": "X1", "name_zh": "=1+1",
                     "quantity": 1, "old_price": None, "new_price": 2.0,
                     "sale_price": None, "urgency": None}],
            "changed": [],
        }
        data = pricing_sheet.build_pricing_xlsx(items, 0.30, "S", "20260529")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=2).value == "X1")
        cell = ws.cell(row=row, column=3)
        # openpyxl 存入 "'=1+1"（带前导撇号），roundtrip 后 data_type='s'（文本）不是 'f'（活公式）
        self.assertNotEqual(cell.data_type, "f")  # 不能是活公式
        self.assertEqual(cell.value, "'=1+1")      # 存储的是带撇号的字面文本


class TestOrchestration(unittest.TestCase):
    def _fixture(self):
        rows = [
            {"barcode": "NEW1", "price": 5.15, "quantity": 100},
            {"barcode": "OLD1", "price": 1.05, "quantity": 50},
            {"barcode": "OLD2", "price": 2.00, "quantity": 10},  # 不变
        ]
        new_entries = [{"barcode": "NEW1", "name": "勺", "invoice_name": "spoon"}]
        products = {
            "OLD1": {"name_zh": "碗", "sale_price": 1.98, "last_purchase_unit_price": 1.10},
            "OLD2": {"name_zh": "铲", "sale_price": 3.0, "last_purchase_unit_price": 2.00},
        }
        summary = [
            {"barcode": "OLD1", "supplier_id": "S1", "margin_pct": 35.0, "urgency_score": 78.0},
            {"barcode": "OLD2", "supplier_id": "S1", "margin_pct": 27.0, "urgency_score": 12.0},
        ]
        p1 = mock.patch.object(pricing_sheet.stockpile_db,
                               "query_products_by_barcodes", return_value=products)
        p2 = mock.patch.object(pricing_sheet.analytics_service,
                               "list_sku_summary", return_value=summary)
        return rows, new_entries, p1, p2

    def test_preview(self):
        rows, new_entries, p1, p2 = self._fixture()
        with p1, p2:
            res = pricing_sheet.preview_pricing(rows, new_entries, "S1")
        self.assertAlmostEqual(res["target_margin_pct"], 31.0)  # median(35,27)
        self.assertEqual(res["n_samples"], 2)
        self.assertEqual(res["n_new"], 1)
        self.assertEqual(res["n_changed"], 1)  # 仅 OLD1 变动
        self.assertEqual(res["skipped_no_baseline"], 0)

    def test_export_bytes_roundtrip(self):
        rows, new_entries, p1, p2 = self._fixture()
        with p1, p2:
            data = pricing_sheet.export_pricing_bytes(
                rows, new_entries, "S1", "雅典XX贸易", 30.0, "20260529")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        self.assertAlmostEqual(ws["B2"].value, 0.30)  # 30.0% → 0.30
        texts = {c.value for r_ in ws.iter_rows() for c in r_ if isinstance(c.value, str)}
        self.assertIn("◆ 新产品", texts)
        self.assertIn("◆ 调价老产品", texts)
