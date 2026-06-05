"""月度总结 Excel 导出契约: build_xlsx + GET /monthly-summary/xlsx/<month>.

6 列明细(供应商/开票日期/总价/税金/特殊税/加税总价) + 末尾合计行, 金额纯数值。
数据隔离: override svc._SUMMARY_DIR 到 tmp(沿用 test_monthly_summary_service 套路)。
"""

import io
import shutil
import unittest
from pathlib import Path

import openpyxl
from flask import Flask

from app.routes.monthly_summary import bp
from app.services import monthly_summary as svc

_TEST_ROOT = Path(__file__).resolve().parent


class _MonthlyXlsxBase(unittest.TestCase):
    def setUp(self):
        self.test_dir = _TEST_ROOT / f"_test_monthly_xlsx_{self._testMethodName}"
        shutil.rmtree(self.test_dir, ignore_errors=True)
        self.test_dir.mkdir(exist_ok=True)
        self._orig_dir = svc._SUMMARY_DIR
        svc._SUMMARY_DIR = self.test_dir

    def tearDown(self):
        svc._SUMMARY_DIR = self._orig_dir
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def _seed_two(self):
        svc.save_record("A贸易", 100.0, 10.0, "2026-04-01", "2026-04", special_tax=5.0)
        svc.save_record("B商行", 200.0, 20.0, "2026-04-02", "2026-04")


class TestBuildXlsx(_MonthlyXlsxBase):
    def test_headers_rows_and_total(self):
        self._seed_two()
        wb = openpyxl.load_workbook(io.BytesIO(svc.build_xlsx("2026-04")))
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["供应商", "开票日期", "总价", "税金", "特殊税", "加税总价"]

        # 两条明细 + 合计行 = 4 行
        assert ws.max_row == 4
        assert ws.cell(row=2, column=1).value == "A贸易"
        assert ws.cell(row=2, column=6).value == 115.0  # 100+10+5
        assert ws.cell(row=3, column=6).value == 220.0  # 200+20

        # 合计行
        assert ws.cell(row=4, column=1).value == "合计"
        assert ws.cell(row=4, column=3).value == 300.0  # 总价
        assert ws.cell(row=4, column=4).value == 30.0  # 税金
        assert ws.cell(row=4, column=5).value == 5.0  # 特殊税
        assert ws.cell(row=4, column=6).value == 335.0  # 加税总价

    def test_amounts_are_numeric_not_string(self):
        self._seed_two()
        ws = openpyxl.load_workbook(io.BytesIO(svc.build_xlsx("2026-04"))).active
        assert isinstance(ws.cell(row=2, column=3).value, (int, float))

    def test_empty_month_has_headers_and_zero_total(self):
        ws = openpyxl.load_workbook(io.BytesIO(svc.build_xlsx("2099-01"))).active
        assert ws.cell(row=1, column=1).value == "供应商"
        # 无明细, 第 2 行即合计
        assert ws.cell(row=2, column=1).value == "合计"
        assert ws.cell(row=2, column=6).value == 0.0


class TestDownloadXlsxRoute(_MonthlyXlsxBase):
    def setUp(self):
        super().setUp()
        self.app = Flask(__name__)
        self.app.register_blueprint(bp)
        self.client = self.app.test_client()

    def test_route_returns_xlsx(self):
        self._seed_two()
        resp = self.client.get("/monthly-summary/xlsx/2026-04")
        assert resp.status_code == 200
        assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ws = openpyxl.load_workbook(io.BytesIO(resp.data)).active
        assert ws.cell(row=1, column=1).value == "供应商"

    def test_route_empty_month_still_200(self):
        resp = self.client.get("/monthly-summary/xlsx/2099-01")
        assert resp.status_code == 200
