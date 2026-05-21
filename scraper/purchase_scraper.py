"""
boson ERP 采购明细抓取 (按月切片 + 缓存 + parquet/xlsx 输出).

跟 sales_scraper.py 的差异 (来自 boson 采购明细页面 F12 抓包, 2026-05-13 确认):
  - range[12]=store_pass_in (采购入库, 实际到货)
    注意: range[0]=stock_order 是 PO 下单台账, 会引入大量 PO 已下未到货幽灵
    记录污染库存事件, 已弃用.
  - event_type = 'purchase'
  - ID号 / 名称 列对应供应商, 不是客户

用法:
  python scraper/purchase_scraper.py                  # 默认最近 1 年
  python scraper/purchase_scraper.py --from 2023-05-12 --to 2026-05-12

输出: SCRAPE_OUTPUT_DIR/events_purchase_<from>_<to>.parquet|xlsx
缓存: SCRAPE_OUTPUT_DIR/_cache/purchase/<begin>_<end>.parquet
"""
import argparse
import os
import re
import sys
import time
from datetime import date
from io import StringIO
from pathlib import Path

import pandas as pd
import requests
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============ 配置 (env var 化) ============
_HERE = Path(__file__).resolve().parent
_REPO_ROOT = _HERE.parent


def _load_env_file(path: Path) -> None:
    """轻量 .env 读取, 不引 python-dotenv 依赖."""
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())


def _resolve(p: str) -> Path:
    path = Path(p)
    return path if path.is_absolute() else _REPO_ROOT / p


_load_env_file(_HERE / ".env")

BOSON_BASE_URL = os.environ.get("BOSON_BASE_URL", "http://bosonapp.local:8137")
URL = f"{BOSON_BASE_URL}/boson/product_search_list.php"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Origin": BOSON_BASE_URL,
    "Referer": URL,
    "Content-Type": "application/x-www-form-urlencoded",
}

_COOKIE_FILE = _resolve(os.environ.get("SCRAPE_COOKIE_FILE", "scraper/cookie.txt"))
if not _COOKIE_FILE.exists():
    sys.exit(
        f"cookie file missing: {_COOKIE_FILE}\n"
        f"create from scraper/cookie.txt.example, 粘贴 PHPSESSID 进去"
    )
_PHPSESSID = _COOKIE_FILE.read_text(encoding="utf-8").strip()
if not _PHPSESSID or _PHPSESSID.startswith("#"):
    sys.exit(f"cookie file is empty or only comments: {_COOKIE_FILE}")

COOKIES = {
    "discount_base_kind": "discount_percent",
    "commission_tax_include_status": "N",
    "PHPSESSID": _PHPSESSID,
}

OUTPUT_DIR = _resolve(os.environ.get("SCRAPE_OUTPUT_DIR", "scraper/staging"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR = OUTPUT_DIR / "_cache" / "purchase"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

DATE_FROM: date = date.today() - relativedelta(years=1)
DATE_TO: date = date.today()
OUTPUT_PARQUET = OUTPUT_DIR / f"events_purchase_{DATE_FROM}_{DATE_TO}.parquet"
OUTPUT_XLSX = OUTPUT_DIR / f"events_purchase_{DATE_FROM}_{DATE_TO}.xlsx"

form_data_template = {
    "root_kind_length": "", "search_time_type": "", "classify_kind_id": "",
    "sign_type": "", "product_kind_id": "", "product_kind_name": "",
    "stack_kind_id": "", "stack_kind_name": "", "product_model": "",
    "document_valid_grade": ">=1", "audit_status": "", "document_kind_id": "",
    "store_id": "", "begin_product_model": "", "end_product_model": "",
    "valid_grade_range": "", "product_batch": "", "valid_grade_symbol": "=",
    "valid_grade": "", "produce_type": "", "product_barcode": "",
    "sequence_number": "", "begin_product_barcode": "", "end_product_barcode": "",
    "client_model": "", "provider_model": "", "begin_unit_price": "",
    "end_unit_price": "", "begin_stockpile_quantity": "",
    "end_stockpile_quantity": "", "product_description": "",
    "description_length": "15", "stockpile_shelf": "", "stockpile_location": "",
    "product_color": "", "product_size": "", "material_description": "",
    "spec_description": "", "document_id": "", "client_document_id": "",
    "product_brand": "", "produce_area": "", "item_remark": "",
    "document_remark": "", "payment_kind_id": "", "source_kind_id": "",
    "industry_kind_id": "", "handler_id": "", "handler_name": "",
    "creator_id": "", "creator_name": "", "web_status_select": "",
    "recommend_status_select": "", "invoice_limit_select": "",
    "discount_limit_select": "", "client_valid_grade_range": "",
    "group_kind_id": "", "money_rate_id": "103", "client_id": "",
    "client_name": "", "client_title": "", "content_type": ".php",
    "language_id": "101",
    "operation_item": "业务明细", "interval_days": "1",
    "document_list_report": "", "time_type": "month", "limit_sum": "",
}

# 采购真值: range[12]=store_pass_in (采购入库)
PURCHASE_RANGES = [
    ("range[12]", "store_pass_in"),
]


# ============ 工具函数 ============
def safe_remove(filepath):
    if not filepath.exists():
        return True
    try:
        filepath.unlink()
        print(f"  已删除旧文件: {filepath.name}")
        return True
    except PermissionError:
        print(f"  ⚠️ {filepath.name} 被占用(可能 Excel 打开),请关闭后重试")
        return False


def fetch_month(begin_str, end_str, range_key="range[12]", range_value="store_pass_in", retry=2):
    form_data = dict(form_data_template)
    form_data[range_key] = range_value
    form_data["begin_date"] = begin_str
    form_data["end_date"] = end_str

    for attempt in range(retry + 1):
        try:
            t0 = time.time()
            r = requests.post(URL, data=form_data, headers=HEADERS,
                              cookies=COOKIES, timeout=600)
            r.encoding = "utf-8"
            size_mb = len(r.content) / 1024 / 1024
            elapsed = time.time() - t0

            if len(r.content) < 50000:
                print(f"    ⚠️ 响应仅 {len(r.content)} 字节,可能 cookie 失效")
                if attempt < retry:
                    time.sleep(5)
                    continue
                return None

            tables = pd.read_html(StringIO(r.text))
            if not tables:
                return None

            df = max(tables, key=len)
            df = df.drop(columns=['Unnamed: 0', 'Unnamed: 1'], errors='ignore')
            print(f"    响应 {size_mb:.1f} MB, 用时 {elapsed:.1f}s, {len(df)} 行")
            return df
        except Exception as e:
            print(f"    ❌ 第 {attempt + 1} 次失败: {e}")
            if attempt < retry:
                time.sleep(5)
    return None


def generate_months(start, end):
    months = []
    cur = start.replace(day=1)
    if cur < start:
        cur = start
    while cur <= end:
        month_end = (cur.replace(day=1) + relativedelta(months=1)) - relativedelta(days=1)
        actual_end = min(month_end, end)
        actual_begin = cur
        months.append((actual_begin.isoformat(), actual_end.isoformat()))
        cur = (cur.replace(day=1) + relativedelta(months=1))
    return months


def clean_raw_dataframe(df):
    int_cols = ['数量', '差数', '等级']
    for col in int_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')

    price_cols = ['单价', '折扣', '金额(€)']
    for col in price_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    numeric_cols = set(int_cols + price_cols)
    for col in df.columns:
        if col not in numeric_cols:
            df[col] = df[col].apply(
                lambda x: '' if pd.isna(x)
                else str(int(x)) if isinstance(x, float) and x.is_integer()
                else str(x).strip()
            )
    return df


# ============ Schema 转换 ============
def extract_category_code(s):
    if not s or pd.isna(s) or s == '':
        return None
    s = str(s).strip()
    m = re.match(r'^([A-Za-z0-9]+(?:-[A-Za-z0-9]+)?)', s)
    return m.group(1) if m else None


def clean_stop_marker(s):
    if s is None or pd.isna(s) or s == '':
        return None
    return re.sub(r'\s*\(停用\)\s*$', '', str(s)).strip() or None


def to_standard_schema(df):
    std = pd.DataFrame()
    std['event_at'] = pd.to_datetime(df['日期'], errors='coerce').dt.strftime('%Y-%m-%d')
    std['event_type'] = 'purchase'
    std['product_barcode'] = df['条形码'].apply(clean_stop_marker).astype('string')
    qty_series = pd.to_numeric(df['数量'], errors='coerce').fillna(0)
    std['qty'] = qty_series.astype('int32')
    std['unit_price'] = pd.to_numeric(df['单价'], errors='coerce').astype('float64')
    std['discount_pct'] = pd.to_numeric(df['折扣'], errors='coerce').fillna(0.0).astype('float64')
    std['document_no'] = df['单号'].astype('string')

    # 采购单: ID号/名称 = 供应商
    std['customer_id'] = pd.Series([None] * len(df), dtype='string')
    std['customer_name'] = pd.Series([None] * len(df), dtype='string')
    std['supplier_id'] = df['ID号'].astype('string')
    std['supplier_name'] = df['名称'].astype('string')

    std['erp_category_raw'] = df['产品种类'].astype('string')
    std['erp_category_code'] = df['产品种类'].apply(extract_category_code).astype('string')
    std['warehouse'] = (df['仓库'].astype('string') if '仓库' in df.columns
                        else pd.Series([None] * len(df), dtype='string'))
    std['product_name_zh'] = (df['品名'].astype('string') if '品名' in df.columns
                              else pd.Series([None] * len(df), dtype='string'))
    std['product_name_local'] = (df['本地品名'].astype('string') if '本地品名' in df.columns
                                 else pd.Series([None] * len(df), dtype='string'))
    std['shipping_doc'] = pd.Series([None] * len(df), dtype='string')

    for col in std.select_dtypes(include='string').columns:
        std[col] = std[col].replace({'': None, 'nan': None, 'None': None, '<NA>': None})

    required = ['event_at', 'product_barcode', 'document_no', 'qty', 'unit_price']
    before = len(std)
    std = std.dropna(subset=required)
    if len(std) < before:
        print(f"    过滤掉 {before - len(std)} 行缺必填字段的数据")
    return std


# ============ 主流程 ============
def main() -> int:
    """跑全流程. 返回非零表示失败 (cookie 失效 / 部分月份抓不到), 给 cron 用."""
    months = generate_months(DATE_FROM, DATE_TO)
    print(f"准备抓取 {len(months)} 个月: {DATE_FROM} → {DATE_TO}")
    print(f"输出 parquet: {OUTPUT_PARQUET}")
    print(f"输出 xlsx:    {OUTPUT_XLSX}")
    print(f"缓存目录:     {CACHE_DIR}")

    print("\n清理旧输出文件...")
    for f in [OUTPUT_PARQUET, OUTPUT_XLSX]:
        if not safe_remove(f):
            print("❌ 旧文件无法删除,关闭后重试")
            return 1
    print()

    total_t0 = time.time()
    all_dfs = []
    failed_months = []

    for i, (begin, end) in enumerate(months, 1):
        cache_file = CACHE_DIR / f"{begin}_{end}.parquet"
        print(f"[{i}/{len(months)}] {begin} → {end}")

        if cache_file.exists():
            df = pd.read_parquet(cache_file)
            print(f"    ✓ 从缓存读取 {len(df)} 行")
            all_dfs.append(df)
            continue

        month_dfs = []
        for range_key, range_value in PURCHASE_RANGES:
            print(f"    抓 {range_key}={range_value}")
            df = fetch_month(begin, end, range_key, range_value)
            if df is not None and len(df) > 0:
                month_dfs.append(df)
            time.sleep(0.3)

        if not month_dfs:
            failed_months.append((begin, end))
            continue

        df = pd.concat(month_dfs, ignore_index=True)
        df = clean_raw_dataframe(df)
        df.to_parquet(cache_file, index=False)
        all_dfs.append(df)
        time.sleep(0.5)

    print(f"\n{'=' * 60}")
    print(f"抓取阶段完成,用时 {(time.time() - total_t0) / 60:.1f} 分钟")
    print(f"成功 {len(all_dfs)}/{len(months)} 个月")
    if failed_months:
        print("失败月份(重跑会自动重试):")
        for b, e in failed_months:
            print(f"  {b} → {e}")

    if not all_dfs:
        print("❌ 全部月份抓取失败 (cookie 可能过期, 去 boson 重登换 PHPSESSID)")
        return 1

    print(f"\n合并 {len(all_dfs)} 个月数据...")
    df = pd.concat(all_dfs, ignore_index=True)
    before = len(df)
    df = df.drop_duplicates()
    print(f"合并 {len(df)} 行 (去重移除 {before - len(df)} 行)")
    df = clean_raw_dataframe(df)

    print("\n转换为标准 schema...")
    t = time.time()
    df_std = to_standard_schema(df)
    print(f"  转换用时 {time.time() - t:.1f}s")
    print(f"  最终 {len(df_std)} 行 × {len(df_std.columns)} 列")

    print(f"\n[1/2] 写 parquet → {OUTPUT_PARQUET.name}")
    t = time.time()
    df_std.to_parquet(OUTPUT_PARQUET, index=False, engine='pyarrow',
                      compression='zstd', compression_level=9)
    pq_size = OUTPUT_PARQUET.stat().st_size / 1024 / 1024
    print(f"    用时 {time.time() - t:.1f}s, 大小 {pq_size:.1f} MB")

    print(f"\n[2/2] 写 xlsx → {OUTPUT_XLSX.name}")
    t = time.time()

    EXCEL_MAX_ROWS = 1048575
    if len(df_std) <= EXCEL_MAX_ROWS:
        df_std.to_excel(OUTPUT_XLSX, index=False, engine="openpyxl")
    else:
        print(f"    ⚠️ {len(df_std)} 行超过单 sheet 上限,分多个 sheet")
        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
            for i in range(0, len(df_std), EXCEL_MAX_ROWS):
                chunk = df_std.iloc[i:i + EXCEL_MAX_ROWS]
                chunk.to_excel(writer, sheet_name=f"data_{i // EXCEL_MAX_ROWS + 1}", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    formats = {
        'event_at': '@', 'product_barcode': '@', 'document_no': '@',
        'customer_id': '@', 'supplier_id': '@', 'erp_category_code': '@',
        'qty': '0', 'unit_price': '0.0000', 'discount_pct': '0.00',
    }
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="E8E8E8")
    header_align = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        header = [cell.value for cell in ws[1]]
        for col_name, fmt in formats.items():
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = fmt
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            sample = list(col_cells)[:200]
            max_len = max((len(str(c.value)) for c in sample if c.value is not None), default=10)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 35)
    wb.save(OUTPUT_XLSX)
    xlsx_size = OUTPUT_XLSX.stat().st_size / 1024 / 1024
    print(f"    用时 {time.time() - t:.1f}s, 大小 {xlsx_size:.1f} MB")

    print(f"\n{'=' * 60}")
    print(f"✅ 完成! 总耗时 {(time.time() - total_t0) / 60:.1f} 分钟")
    print(f"   parquet: {OUTPUT_PARQUET}")
    print(f"   xlsx:    {OUTPUT_XLSX}")

    if failed_months:
        print(f"\n⚠️  有 {len(failed_months)} 个月份失败 (见上面列表), 返回非零")
        return 1
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="boson ERP 采购明细抓取 (按月切 + 缓存 + parquet/xlsx)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="cookie 失效时去 scraper/cookie.txt 换 PHPSESSID",
    )
    _today = date.today()
    parser.add_argument("--from", dest="date_from", type=date.fromisoformat,
                        default=_today - relativedelta(years=1),
                        help="起始日期 YYYY-MM-DD (默认: 今天往前 1 年)")
    parser.add_argument("--to", dest="date_to", type=date.fromisoformat,
                        default=_today,
                        help="结束日期 YYYY-MM-DD (默认: 今天)")
    args = parser.parse_args()
    if args.date_from > args.date_to:
        parser.error(f"--from ({args.date_from}) 不能晚于 --to ({args.date_to})")

    DATE_FROM = args.date_from
    DATE_TO = args.date_to
    OUTPUT_PARQUET = OUTPUT_DIR / f"events_purchase_{DATE_FROM}_{DATE_TO}.parquet"
    OUTPUT_XLSX = OUTPUT_DIR / f"events_purchase_{DATE_FROM}_{DATE_TO}.xlsx"

    sys.exit(main())
