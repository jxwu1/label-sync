"""采购定价表（老板报价表）服务。

纯函数核心（中位数 / 目标利润率 / 调价检测 / 条目组装 / xlsx 生成）+ 薄编排层。
设计文档：docs/superpowers/specs/2026-05-29-采购定价表-design.md
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.repositories import stockpile_db
from app.services import analytics as analytics_service

# xlsx 列序（1-indexed）
_COL_IMG, _COL_BARCODE, _COL_NAME, _COL_QTY = 1, 2, 3, 4
_COL_OLD, _COL_NEW, _COL_SALE, _COL_CUR_MARGIN = 5, 6, 7, 8
_COL_SUGGEST, _COL_SUGGEST_MARGIN, _COL_EDIT, _COL_EDIT_MARGIN, _COL_HEAT = 9, 10, 11, 12, 13
_LAST_COL = _COL_HEAT

_EUR_FMT = '#,##0.00 "€"'
_PCT_FMT = "0.00%"
_TARGET_CELL = "$B$2"
_PRICE_DECIMALS = 4


def _median(values: list[float | None]) -> float | None:
    nums = sorted(v for v in values if v is not None)
    n = len(nums)
    if n == 0:
        return None
    mid = n // 2
    if n % 2 == 1:
        return float(nums[mid])
    return (nums[mid - 1] + nums[mid]) / 2.0


def compute_target_margin_pct(summary_items: list[dict], supplier_id: str) -> dict:
    """该供应商在售老品 margin_pct（百分数）的中位数。无供应商/无数据 → median=None。"""
    if not supplier_id:
        return {"median": None, "n_samples": 0}
    vals = [
        it["margin_pct"]
        for it in summary_items
        if it.get("supplier_id") == supplier_id and it.get("margin_pct") is not None
    ]
    return {"median": _median(vals), "n_samples": len(vals)}


def _prices_equal(a: float, b: float) -> bool:
    return round(a, _PRICE_DECIMALS) == round(b, _PRICE_DECIMALS)


def detect_changed_old(rows: list[dict], baselines: dict[str, float | None]) -> tuple[list[dict], int]:
    """上传行中、在 baselines（=主档已知）里且进价 ≠ 最近采购净价的老品。

    baselines: {barcode: last_purchase_unit_price|None}。值为 None=无基准→跳过并计数。
    返回 (changed_list, skipped_no_baseline)。
    """
    changed: list[dict] = []
    skipped_no_baseline = 0
    seen: set[str] = set()
    for r in rows:
        bc = r["barcode"]
        if bc in seen or bc not in baselines:
            continue
        seen.add(bc)
        baseline = baselines[bc]
        if baseline is None:
            skipped_no_baseline += 1
            continue
        new_price = float(r["price"])
        if not _prices_equal(new_price, float(baseline)):
            changed.append({
                "barcode": bc,
                "old_price": float(baseline),
                "new_price": new_price,
                "quantity": int(r["quantity"]),
            })
    return changed, skipped_no_baseline


def build_pricing_items(
    new_entries: list[dict],
    rows: list[dict],
    products_by_barcode: dict[str, dict],
    summary_by_barcode: dict[str, dict],
) -> dict:
    """组装定价表条目。

    new_entries: [{barcode, name, invoice_name}]（新品，名取这里）
    rows: [{barcode, price, quantity}]（上传行，取价/量）
    products_by_barcode: 主档已知产品 {bc: {name_zh, sale_price, last_purchase_unit_price}}
    summary_by_barcode: {bc: list_sku_summary item}（取 urgency_score）
    """
    row_by_bc: dict[str, dict] = {}
    for r in rows:
        bc = r["barcode"]
        if bc not in row_by_bc:
            row_by_bc[bc] = {"price": float(r["price"]), "quantity": int(r["quantity"])}

    new_items: list[dict] = []
    seen_new: set[str] = set()
    for e in new_entries:
        bc = e["barcode"]
        if bc in seen_new:
            continue
        seen_new.add(bc)
        rr = row_by_bc.get(bc, {"price": 0.0, "quantity": 0})
        new_items.append({
            "section": "new", "barcode": bc, "name_zh": e.get("name") or "",
            "quantity": rr["quantity"], "old_price": None, "new_price": rr["price"],
            "sale_price": None, "urgency": None,
        })

    baselines = {bc: p.get("last_purchase_unit_price") for bc, p in products_by_barcode.items()}
    changed_raw, skipped = detect_changed_old(rows, baselines)
    changed_items: list[dict] = []
    for c in changed_raw:
        bc = c["barcode"]
        prod = products_by_barcode.get(bc, {})
        summ = summary_by_barcode.get(bc, {})
        changed_items.append({
            "section": "changed", "barcode": bc, "name_zh": prod.get("name_zh") or "",
            "quantity": c["quantity"], "old_price": c["old_price"], "new_price": c["new_price"],
            "sale_price": prod.get("sale_price"), "urgency": summ.get("urgency_score"),
        })

    return {"new": new_items, "changed": changed_items, "skipped_no_baseline": skipped}


_HEADERS = [
    "图片", "条码", "中文品名", "数量", "旧进价", "新进价", "现售价",
    "现利润率", "建议批发价", "建议利润率", "修改", "改后利润率", "热度",
]


def _write_item_row(ws, r: int, item: dict) -> None:
    is_new = item["section"] == "new"
    ws.cell(row=r, column=_COL_BARCODE, value=item["barcode"]).number_format = "@"
    _name = item["name_zh"]
    if _name and _name[0] in ("=", "+", "-", "@"):
        _name = "'" + _name  # 中和公式注入：以公式触发符开头的品名强制为文本
    ws.cell(row=r, column=_COL_NAME, value=_name)
    ws.cell(row=r, column=_COL_QTY, value=item["quantity"]).number_format = "0"
    if item["old_price"] is not None:
        ws.cell(row=r, column=_COL_OLD, value=item["old_price"]).number_format = _EUR_FMT
    ws.cell(row=r, column=_COL_NEW, value=item["new_price"]).number_format = _EUR_FMT
    if item["sale_price"] is not None:
        ws.cell(row=r, column=_COL_SALE, value=item["sale_price"]).number_format = _EUR_FMT
    # 现利润率（仅老品有现售价时）：(现售价-新进价)/现售价
    if not is_new:
        ws.cell(row=r, column=_COL_CUR_MARGIN,
                value=f'=IF(AND(G{r},F{r}),(G{r}-F{r})/G{r},"")').number_format = _PCT_FMT
    # 建议批发价 = 新进价/(1-目标利润率)；建议利润率 = (建议-新进价)/建议
    ws.cell(row=r, column=_COL_SUGGEST, value=f"=F{r}/(1-{_TARGET_CELL})").number_format = _EUR_FMT
    ws.cell(row=r, column=_COL_SUGGEST_MARGIN,
            value=f'=IF(I{r},(I{r}-F{r})/I{r},"")').number_format = _PCT_FMT
    # 修改（空，老板填）+ 改后利润率
    ws.cell(row=r, column=_COL_EDIT).number_format = _EUR_FMT
    ws.cell(row=r, column=_COL_EDIT_MARGIN,
            value=f'=IF(K{r},(K{r}-F{r})/K{r},"")').number_format = _PCT_FMT
    # 热度
    if is_new:
        ws.cell(row=r, column=_COL_HEAT, value="新品")
    elif item["urgency"] is not None:
        ws.cell(row=r, column=_COL_HEAT, value=item["urgency"]).number_format = "0"
    ws.row_dimensions[r].height = 90  # 留图片高度


def _write_section_header(ws, r: int, title: str, fill: PatternFill) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_LAST_COL)
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = Font(bold=True)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")


def build_pricing_xlsx(items: dict, target_fraction: float, supplier_name: str, date_str: str) -> bytes:
    """生成定价表 xlsx（bytes）。target_fraction 是小数（0.30），写入 B2 并被公式引用。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "定价表"

    # 顶部信息块
    ws["A1"] = "供应商:"
    ws["B1"] = supplier_name
    ws["D1"] = "日期:"
    ws["E1"] = date_str
    ws["A2"] = "目标利润率:"
    ws["B2"] = float(target_fraction)
    ws["B2"].number_format = _PCT_FMT
    ws["D2"] = "说明:"
    ws["E2"] = '改"修改"列利润率自动重算'

    # 表头（第 4 行）
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="E8E8E8")
    header_align = Alignment(horizontal="center", vertical="center")
    header_row = 4
    for idx, name in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    section_fill = PatternFill("solid", fgColor="DCEBFA")
    r = header_row + 1
    if items["new"]:
        _write_section_header(ws, r, "◆ 新产品", section_fill)
        r += 1
        for it in items["new"]:
            _write_item_row(ws, r, it)
            r += 1
    if items["changed"]:
        _write_section_header(ws, r, "◆ 调价老产品", section_fill)
        r += 1
        for it in items["changed"]:
            _write_item_row(ws, r, it)
            r += 1

    # 列宽 + 冻结表头
    widths = {1: 18, 2: 16, 3: 22, 4: 6, 5: 10, 6: 10, 7: 10, 8: 9, 9: 11, 10: 9, 11: 10, 12: 10, 13: 7}
    for col_idx, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gather(rows: list[dict], new_entries: list[dict], supplier_id: str):
    uploaded_bcs = [r["barcode"] for r in rows]
    products = stockpile_db.query_products_by_barcodes(uploaded_bcs)
    summary = analytics_service.list_sku_summary()
    summary_by_bc = {it["barcode"]: it for it in summary}
    target = compute_target_margin_pct(summary, supplier_id)
    items = build_pricing_items(new_entries, rows, products, summary_by_bc)
    return target, items


def preview_pricing(rows: list[dict], new_entries: list[dict], supplier_id: str) -> dict:
    target, items = _gather(rows, new_entries, supplier_id)
    return {
        "target_margin_pct": target["median"],
        "n_samples": target["n_samples"],
        "n_new": len(items["new"]),
        "n_changed": len(items["changed"]),
        "skipped_no_baseline": items["skipped_no_baseline"],
    }


def export_pricing_bytes(
    rows: list[dict], new_entries: list[dict], supplier_id: str,
    supplier_name: str, target_margin_pct: float, date_str: str,
) -> bytes:
    _, items = _gather(rows, new_entries, supplier_id)
    target_fraction = float(target_margin_pct) / 100.0
    return build_pricing_xlsx(items, target_fraction, supplier_name, date_str)
