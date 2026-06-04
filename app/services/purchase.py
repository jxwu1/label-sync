from app.config import CONFIG
import csv
import io
import math
import zipfile
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
import pandas as pd

from app.repositories import stockpile_db

_TEMPLATE_PATH = CONFIG.resource_dir / "static" / "templates" / "产品信息导入模板.csv"

# 供应商 Excel 列索引（0-based）
_SUPPLIER_BARCODE_COL = 0
_SUPPLIER_PRICE_COL = 2
_SUPPLIER_QUANTITY_COL = 5

# 产品信息导入模板列索引（同一字段需要写入多列）
_TEMPLATE_BARCODE_COLS = (0, 1, 10)
_TEMPLATE_NAME_COL = 3
_TEMPLATE_INVOICE_NAME_COL = 4
_TEMPLATE_ENGLISH_NAME_COL = 5  # 第6列(英语品名): 按业务要求也填发票品名
_TEMPLATE_SUPPLIER_ID_COL = 38
_TEMPLATE_SUPPLIER_NAME_COL = 39

_PRICE_DECIMALS = 4
_PRICE_QUANT = Decimal("0.0001")


def _round_half_up(price: float) -> float:
    return float(Decimal(str(price)).quantize(_PRICE_QUANT, rounding=ROUND_HALF_UP))


def _parse_price(raw) -> tuple[float, bool]:
    """返回 (价格, 是否需要人工复核)。超过4位自动四舍五入；NaN/无法解析才 flagged。"""
    try:
        price = float(raw)
    except (ValueError, TypeError):
        return 0.0, True
    if math.isnan(price) or math.isinf(price):
        return 0.0, True
    return _round_half_up(price), False


def _parse_quantity(raw) -> tuple[int, bool]:
    """返回 (数量, 是否可疑)。无法解析 → (0, True); <=0 → (值, True)。

    可疑标记给前端高亮, 避免数量打错被静默当 0 导致订货短缺。
    """
    try:
        q = int(float(raw))
    except (ValueError, TypeError):
        return 0, True
    return q, (q <= 0)


@dataclass
class PurchaseRow:
    barcode: str
    price_raw: str
    price: float
    quantity: int
    price_flagged: bool
    quantity_flagged: bool = False

    def formatted(self) -> str:
        return f"{self.barcode},{self.price:.4f},,{self.quantity}"

    def to_dict(self) -> dict:
        return {
            "barcode": self.barcode,
            "price": self.price,
            "quantity": self.quantity,
            "price_flagged": self.price_flagged,
            "quantity_flagged": self.quantity_flagged,
            "formatted": self.formatted(),
        }


def parse_purchase_excel(file_bytes: bytes) -> list[PurchaseRow]:
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=0, dtype=str, engine="calamine")
    except Exception as exc:
        raise ValueError("无法读取供应商 Excel 文件，请确认上传的是有效的 .xlsx 文件") from exc
    rows = []
    for _, row in df.iterrows():
        try:
            barcode = str(row.iloc[_SUPPLIER_BARCODE_COL]).strip()
            price_val = row.iloc[_SUPPLIER_PRICE_COL]
            qty_val = row.iloc[_SUPPLIER_QUANTITY_COL]
        except IndexError as exc:
            raise ValueError("供应商 Excel 列数不足，应至少含 条码 / 价格 / 数量 列") from exc
        price, price_flagged = _parse_price(price_val)
        quantity, quantity_flagged = _parse_quantity(qty_val)
        rows.append(
            PurchaseRow(
                barcode=barcode,
                price_raw=str(price_val),
                price=price,
                quantity=quantity,
                price_flagged=price_flagged,
                quantity_flagged=quantity_flagged,
            )
        )
    return rows


def build_output_excel(file_bytes: bytes, rows_data: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value="导入信息")
    for i, row in enumerate(rows_data):
        ws.cell(row=i + 2, column=col, value=row["formatted"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_new_barcodes(entries: list[dict]) -> int:
    count = 0
    for entry in entries:
        barcode = entry.get("barcode", "").strip()
        if not barcode:
            continue
        stockpile_db.insert_or_update(
            barcode=barcode,
            model=barcode,
            location="",
            source="purchase_import",
        )
        count += 1
    return count


def find_new_barcodes(rows: list[PurchaseRow], system_set: set[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for r in rows:
        bc = r.barcode
        if bc in system_set or bc in seen:
            continue
        seen.add(bc)
        out.append(bc)
    return out


def lookup_dominant_supplier(barcodes: list[str]) -> dict | None:
    """按「每个货号最近一次采购的供应商」统计，返回出现最多的那家。

    业务前提：一份供应商 Excel 对应一家供应商，所以多数决足够；
    全部货号都查不到供应商（如全新条码）时返回 None。

    返回 {"supplier_id": str, "supplier_name": str, "matched": int, "total": int}。
    """
    from collections import Counter

    from sqlalchemy import select

    from app.models import InventoryEvent, Supplier

    unique_bcs = sorted({bc for bc in barcodes if bc})
    if not unique_bcs:
        return None

    counter: Counter[tuple[str, str]] = Counter()
    with stockpile_db._session() as session:
        for bc in unique_bcs:
            row = session.execute(
                select(InventoryEvent.supplier_id, Supplier.supplier_name)
                .join(Supplier, Supplier.supplier_id == InventoryEvent.supplier_id)
                .where(
                    InventoryEvent.product_barcode == bc,
                    InventoryEvent.event_type == "purchase",
                    InventoryEvent.supplier_id.isnot(None),
                )
                .order_by(InventoryEvent.event_at.desc())
                .limit(1)
            ).first()
            if row is None:
                continue
            counter[(row.supplier_id, row.supplier_name)] += 1

    if not counter:
        return None
    (sid, sname), matched = counter.most_common(1)[0]
    return {
        "supplier_id": sid,
        "supplier_name": sname,
        "matched": matched,
        "total": len(unique_bcs),
    }


def _read_template_header() -> list[str]:
    with _TEMPLATE_PATH.open("rb") as f:
        raw = f.read()
    try:
        text = raw.decode("gbk")
    except UnicodeDecodeError:
        text = raw.decode("utf-8")
    first_line = text.splitlines()[0]
    return next(csv.reader([first_line]))


def _fill_template_row(entry: dict, n_cols: int) -> list[str]:
    row = [""] * n_cols
    for i in _TEMPLATE_BARCODE_COLS:
        if i < n_cols:
            row[i] = entry["barcode"]
    if _TEMPLATE_NAME_COL < n_cols:
        row[_TEMPLATE_NAME_COL] = entry["name"]
    if _TEMPLATE_INVOICE_NAME_COL < n_cols:
        row[_TEMPLATE_INVOICE_NAME_COL] = entry["invoice_name"]
    if _TEMPLATE_ENGLISH_NAME_COL < n_cols:
        row[_TEMPLATE_ENGLISH_NAME_COL] = entry["invoice_name"]
    if _TEMPLATE_SUPPLIER_ID_COL < n_cols:
        row[_TEMPLATE_SUPPLIER_ID_COL] = entry["supplier_id"]
    if _TEMPLATE_SUPPLIER_NAME_COL < n_cols:
        row[_TEMPLATE_SUPPLIER_NAME_COL] = entry["supplier_name"]
    return row


def build_template_csv(new_entries: list[dict]) -> bytes:
    header = _read_template_header()
    n_cols = len(header)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for entry in new_entries:
        writer.writerow(_fill_template_row(entry, n_cols))
    return buf.getvalue().encode("gbk")


def build_zip(purchase_xlsx: bytes, template_csv: bytes | None, date_str: str) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"采购订单{date_str}.xlsx", purchase_xlsx)
        if template_csv is not None:
            zf.writestr("产品信息导入模板.csv", template_csv)
    return buf.getvalue()


# ── Purchase Order tracking ─────────────────────────────────────────


def create_order(
    rows: list[PurchaseRow],
    *,
    supplier_id: str | None = None,
    source_file: str | None = None,
) -> dict:
    from datetime import date as date_cls
    from app.models import PurchaseOrder, PurchaseOrderLine, get_session

    from sqlalchemy import select

    total_qty = sum(r.quantity for r in rows)
    total_amount = round(sum(r.price * r.quantity for r in rows), 2)

    with get_session() as s:
        # 幂等防重复: 同文件 + 同总量 + 同总额 且仍未到货(placed) → 视为重复上传,
        # 返回已有单而非再建一张. 已到货的旧单不拦(那是新一批补货).
        if source_file:
            existing = s.execute(
                select(PurchaseOrder)
                .where(
                    PurchaseOrder.source_file == source_file,
                    PurchaseOrder.status == "placed",
                    PurchaseOrder.total_qty == total_qty,
                    PurchaseOrder.total_amount == total_amount,
                )
                .order_by(PurchaseOrder.id.desc())
                .limit(1)
            ).scalar_one_or_none()
            if existing is not None:
                return {
                    "order_id": existing.id,
                    "total_qty": existing.total_qty,
                    "total_amount": existing.total_amount,
                    "duplicate": True,
                }

    with get_session() as s:
        order = PurchaseOrder(
            supplier_id=supplier_id,
            order_date=date_cls.today().isoformat(),
            status="placed",
            source_file=source_file,
            total_qty=total_qty,
            total_amount=round(total_amount, 2),
        )
        s.add(order)
        s.flush()
        for r in rows:
            s.add(
                PurchaseOrderLine(
                    order_id=order.id,
                    product_barcode=r.barcode,
                    qty_ordered=r.quantity,
                    unit_price=r.price,
                )
            )
        order_id = order.id

    return {
        "order_id": order_id,
        "total_qty": total_qty,
        "total_amount": total_amount,
        "duplicate": False,
    }


def list_orders(limit: int = 50) -> list[dict]:
    from sqlalchemy import select
    from app.models import PurchaseOrder, Supplier, get_session

    with get_session() as s:
        rows = s.execute(
            select(PurchaseOrder, Supplier.supplier_name)
            .outerjoin(Supplier, Supplier.supplier_id == PurchaseOrder.supplier_id)
            .order_by(PurchaseOrder.id.desc())
            .limit(limit)
        ).all()
        return [
            {
                "id": po.id,
                "supplier_id": po.supplier_id,
                "supplier_name": sname or "",
                "order_date": po.order_date,
                "arrival_date": po.arrival_date,
                "status": po.status,
                "source_file": po.source_file,
                "total_qty": po.total_qty,
                "total_amount": po.total_amount,
            }
            for po, sname in rows
        ]


def record_arrival(order_id: int, arrival_date: str) -> dict:
    from sqlalchemy import select
    from app.models import PurchaseOrder, PurchaseOrderLine, get_session

    with get_session() as s:
        order = s.get(PurchaseOrder, order_id)
        if not order:
            raise ValueError(f"订单不存在：{order_id}")
        order.arrival_date = arrival_date
        order.status = "arrived"
        lines = (
            s.execute(select(PurchaseOrderLine).where(PurchaseOrderLine.order_id == order_id))
            .scalars()
            .all()
        )
        for line in lines:
            line.qty_arrived = line.qty_ordered

    return {"order_id": order_id, "arrival_date": arrival_date, "status": "arrived"}


def compute_supplier_lead_times(limit: int = 50) -> list[dict]:
    from statistics import median
    from sqlalchemy import select
    from datetime import date as date_cls
    from app.models import PurchaseOrder, Supplier, get_session

    with get_session() as s:
        suppliers = s.execute(
            select(PurchaseOrder.supplier_id, Supplier.supplier_name)
            .join(Supplier, Supplier.supplier_id == PurchaseOrder.supplier_id)
            .where(PurchaseOrder.arrival_date.isnot(None))
            .group_by(PurchaseOrder.supplier_id)
        ).all()

        results = []
        for sid, sname in suppliers:
            orders = s.execute(
                select(PurchaseOrder.order_date, PurchaseOrder.arrival_date)
                .where(PurchaseOrder.supplier_id == sid, PurchaseOrder.arrival_date.isnot(None))
                .order_by(PurchaseOrder.order_date.desc())
                .limit(limit)
            ).all()
            lts = []
            for o_date, a_date in orders:
                try:
                    d = (date_cls.fromisoformat(a_date) - date_cls.fromisoformat(o_date)).days
                    if d >= 0:
                        lts.append(d)
                except (ValueError, TypeError):
                    continue
            if lts:
                results.append(
                    {
                        "supplier_id": sid,
                        "supplier_name": sname,
                        "n_samples": len(lts),
                        "median_days": median(lts),
                        "min_days": min(lts),
                        "max_days": max(lts),
                    }
                )
        return results
