import csv
import io
import math
import zipfile
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
import pandas as pd

import stockpile_db

_TEMPLATE_PATH = Path(__file__).resolve().parent / "static" / "templates" / "产品信息导入模板.csv"

# 供应商 Excel 列索引（0-based）
_SUPPLIER_BARCODE_COL = 0
_SUPPLIER_PRICE_COL = 2
_SUPPLIER_QUANTITY_COL = 5

# 产品信息导入模板列索引（同一字段需要写入多列）
_TEMPLATE_BARCODE_COLS = (0, 1, 10)
_TEMPLATE_NAME_COL = 3
_TEMPLATE_INVOICE_NAME_COL = 4
_TEMPLATE_SUPPLIER_ID_COL = 38
_TEMPLATE_SUPPLIER_NAME_COL = 39

_PRICE_DECIMALS = 4
_PRICE_QUANT = Decimal("0.0001")


def _round_half_up(price: float) -> float:
    return float(Decimal(str(price)).quantize(_PRICE_QUANT, rounding=ROUND_HALF_UP))


def _parse_price(raw) -> tuple[float, bool]:
    """返回 (价格, 是否需要人工复核)。超过4位自动四舍五入；NaN/无法解析才 flagged。"""
    try:
        price = float(raw)
    except (ValueError, TypeError):
        return 0.0, True
    if math.isnan(price) or math.isinf(price):
        return 0.0, True
    return _round_half_up(price), False


def _parse_quantity(raw) -> int:
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return 0


@dataclass
class PurchaseRow:
    barcode: str
    price_raw: str
    price: float
    quantity: int
    price_flagged: bool

    def formatted(self) -> str:
        return f"{self.barcode},{self.price:.4f},,{self.quantity}"

    def to_dict(self) -> dict:
        return {
            "barcode": self.barcode,
            "price": self.price,
            "quantity": self.quantity,
            "price_flagged": self.price_flagged,
            "formatted": self.formatted(),
        }


def parse_purchase_excel(file_bytes: bytes) -> list[PurchaseRow]:
    df = pd.read_excel(io.BytesIO(file_bytes), header=0, dtype=str, engine="calamine")
    rows = []
    for _, row in df.iterrows():
        barcode = str(row.iloc[_SUPPLIER_BARCODE_COL]).strip()
        price_val = row.iloc[_SUPPLIER_PRICE_COL]
        qty_val = row.iloc[_SUPPLIER_QUANTITY_COL]
        price, price_flagged = _parse_price(price_val)
        rows.append(
            PurchaseRow(
                barcode=barcode,
                price_raw=str(price_val),
                price=price,
                quantity=_parse_quantity(qty_val),
                price_flagged=price_flagged,
            )
        )
    return rows


def build_output_excel(file_bytes: bytes, rows_data: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value="导入信息")
    for i, row in enumerate(rows_data):
        ws.cell(row=i + 2, column=col, value=row["formatted"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_new_barcodes(entries: list[dict]) -> int:
    count = 0
    for entry in entries:
        barcode = entry.get("barcode", "").strip()
        if not barcode:
            continue
        stockpile_db.insert_or_update(
            barcode=barcode,
            model=barcode,
            location="",
            source="purchase_import",
        )
        count += 1
    return count


def find_new_barcodes(rows: list[PurchaseRow], system_set: set[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for r in rows:
        bc = r.barcode
        if bc in system_set or bc in seen:
            continue
        seen.add(bc)
        out.append(bc)
    return out


def _read_template_header() -> list[str]:
    with _TEMPLATE_PATH.open("rb") as f:
        raw = f.read()
    try:
        text = raw.decode("gbk")
    except UnicodeDecodeError:
        text = raw.decode("utf-8")
    first_line = text.splitlines()[0]
    return next(csv.reader([first_line]))


def _fill_template_row(entry: dict, n_cols: int) -> list[str]:
    row = [""] * n_cols
    for i in _TEMPLATE_BARCODE_COLS:
        if i < n_cols:
            row[i] = entry["barcode"]
    if _TEMPLATE_NAME_COL < n_cols:
        row[_TEMPLATE_NAME_COL] = entry["name"]
    if _TEMPLATE_INVOICE_NAME_COL < n_cols:
        row[_TEMPLATE_INVOICE_NAME_COL] = entry["invoice_name"]
    if _TEMPLATE_SUPPLIER_ID_COL < n_cols:
        row[_TEMPLATE_SUPPLIER_ID_COL] = entry["supplier_id"]
    if _TEMPLATE_SUPPLIER_NAME_COL < n_cols:
        row[_TEMPLATE_SUPPLIER_NAME_COL] = entry["supplier_name"]
    return row


def build_template_csv(new_entries: list[dict]) -> bytes:
    header = _read_template_header()
    n_cols = len(header)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for entry in new_entries:
        writer.writerow(_fill_template_row(entry, n_cols))
    return buf.getvalue().encode("gbk")


def build_zip(purchase_xlsx: bytes, template_csv: bytes | None, date_str: str) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"采购订单{date_str}.xlsx", purchase_xlsx)
        if template_csv is not None:
            zf.writestr("产品信息导入模板.csv", template_csv)
    return buf.getvalue()
